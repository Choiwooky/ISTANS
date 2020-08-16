import pandas as pd
import openpyxl
import os

path = './RawData'
file_list = os.listdir(path)
print("file_list : {}". format(file_list))

 year = ['15', '16', '17', '18', '19', '20']
# year = ['19']
 month = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
# month = ['04', '05', '06', '07', '08', '09', '10', '11', '12']


# 통계표_시군구
for y in year:
    for m in month:
        file_dir = 'RawData/20'+y+'년_'+m+'월_자동차_등록자료_통계.xlsx'
        filename = '20'+y+'년_'+m+'월_자동차_등록자료_통계.xlsx'
        if filename not in file_list:
            continue
        wb = openpyxl.load_workbook(file_dir)
        sheets = wb.sheetnames
        target = [x for x in sheets if '통계표_시군구' in x]

        for s in sheets:
            if s not in target:
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
        Sfilename = 'CleaningData/20'+y+'년'+m+'월_통계표_시군구.xlsx'
        wb.save(Sfilename)

# 연료별_등록현황
for y in year:
    for m in month:
        file_dir = 'RawData/20'+y+'년_'+m+'월_자동차_등록자료_통계.xlsx'
        filename = '20'+y+'년_'+m+'월_자동차_등록자료_통계.xlsx'
        if filename not in file_list:
            continue
        wb = openpyxl.load_workbook(file_dir)
        sheets = wb.sheetnames
        target = [x for x in sheets if '연료별_등록현황' in x]

        for s in sheets:
            if s not in target:
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
        Sfilename = '수소,전기/20'+y+'년'+m+'월_연료별_등록현황.xlsx'
        wb.save(Sfilename)

# 인구 자료 전처리
df = pd.read_csv('통계청 시군구 총 인구수_KOSIS(전처리).csv', encoding='cp949')
for i, z in enumerate(df['시군구']):
    if z[-1] == '시' or z[-1] == '군' or z[-1] == '구':
        df.at[i, '시군구'] = z[:-1]
    else:
        df.at[i, '시군구'] =  z
df.to_csv('./Result/인구자료.csv', encoding='cp949')
